from telegram import Update, InputFile
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, CallbackContext
import yt_dlp
import os
import subprocess
import requests
from bs4 import BeautifulSoup
import re
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook

TOKEN_FILE = 'tokens.json'
API_TOKEN = ""
GENIUS_TOKEN = ""

# --- УМНАЯ ПРОВЕРКА ТОКЕНОВ ПРИ ЗАПУСКЕ ---
if os.path.exists(TOKEN_FILE):
    try:
        with open(TOKEN_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            API_TOKEN = data.get('API_TOKEN', '')
            GENIUS_TOKEN = data.get('GENIUS_TOKEN', '')
        print("--- СЕРВЕР ПЕРЕЗАГРУЖЕН: ТОКЕНЫ УСПЕШНО ВОССТАНОВЛЕНЫ ИЗ ПАМЯТИ ---")
    except Exception:
        pass

# Если токенов нет в файле, запрашиваем их вручную
if not API_TOKEN:
    print("--- НАСТРОЙКА БОТА ---")
    API_TOKEN = input("Введите ваш API от Telegram: ").strip()
    while not API_TOKEN:
        print("Без токена Telegram бот не сможет работать!")
        API_TOKEN = input("Введите ваш API от Telegram: ").strip()

    GENIUS_TOKEN = input("Введите ваш API от Genius (нажмите Enter, чтобы пропустить): ").strip()
    
    # Сохраняем данные, чтобы бот "вспомнил" их при перезагрузке сервера
    try:
        with open(TOKEN_FILE, 'w', encoding='utf-8') as f:
            json.dump({'API_TOKEN': API_TOKEN, 'GENIUS_TOKEN': GENIUS_TOKEN}, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"Не удалось сохранить токены в файл: {e}")

print("----------------------------------------------------------------")
print("🚀 БОТ УСПЕШНО ЗАПУЩЕН И ГОТОВ К РАБОТЕ!")
print("👉 Чтобы закрыть терминал и оставить бота работать в фоне,")
print("   нажмите комбинацию клавиш: Ctrl + P, затем Ctrl + Q")
print("----------------------------------------------------------------")

LOG_FILE = 'music_history.xlsx'
BANNED_USERS = [718895955, 6181518024] 
user_audio = {}

def start(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id
    if user_id in BANNED_USERS:
        return

    welcome_text = (
        "🎵 Привет! Музыкальный бот на связи! 🎶\n\n"
        "Просто напиши исполнителя и название песни, и я постараюсь найти её для тебя! "
        "Например: <b>Король и шут лесник</b>\n\n"
        "Желаю приятного прослушивания! 😊"
    )
    update.message.reply_text(welcome_text, parse_mode='HTML')

def handle_message(update: Update, context: CallbackContext):
    user = update.message.from_user
    user_id = user.id
    
    if user_id in BANNED_USERS:
        return

    if user_id in user_audio:
        text = update.message.text.strip()
        if '-' not in text:
            update.message.reply_text("Укажи промежуток в формате 68-73")
            return

        try:
            start_sec, end_sec = map(int, text.split('-'))
            if start_sec >= end_sec:
                update.message.reply_text("Ошибка: начало должно быть меньше конца.")
                return
        except Exception:
            update.message.reply_text("Ошибка: укажи промежуток в формате 68-73")
            return

        input_file, orig_name = user_audio.pop(user_id)
        output_file = f"cut_{input_file}"

        cmd = [
            'ffmpeg', '-y', '-i', input_file,
            '-ss', str(start_sec), '-to', str(end_sec),
            '-c', 'copy', output_file
        ]
        subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

        if os.path.exists(output_file):
            base_name = os.path.splitext(orig_name)[0] if orig_name else "audio"
            new_name = f"{base_name} ({start_sec}-{end_sec} сек).mp3"
            with open(output_file, 'rb') as audio:
                update.message.reply_audio(
                    InputFile(audio, filename=new_name),
                    title=new_name
                )
            os.remove(output_file)
        else:
            update.message.reply_text("Не удалось вырезать фрагмент.")

        os.remove(input_file)
        return

    query = update.message.text
    
    username = f"@{user.username}" if user.username else user.first_name
    save_query_to_excel(user_id, username, query)

    searching_msg = update.message.reply_text("🔍 Ищу и скачиваю песню, подожди немного...")
    track = search_youtube_track(query)
    if track:
        title, url = track
        filename = download_audio(url)
        if filename:
            try:
                context.bot.delete_message(chat_id=update.message.chat_id, message_id=searching_msg.message_id)
            except Exception:
                pass
            with open(filename, 'rb') as audio:
                update.message.reply_audio(audio, title=title, timeout=120)
            os.remove(filename)
            sponsor_text = "🤖 Спонсор: @QuantumSageBot — ИИ ассистент нового поколения!"
            update.message.reply_text(sponsor_text)
            
            if GENIUS_TOKEN:
                lyrics_search_msg = update.message.reply_text("Попробую найти текст песни")
                lyrics = get_lyrics_from_genius_api(query)
                try:
                    context.bot.delete_message(chat_id=update.message.chat_id, message_id=lyrics_search_msg.message_id)
                except Exception:
                    pass
                if lyrics:
                    update.message.reply_text(f"📝 Текст песни:\n\n{lyrics}")
                else:
                    update.message.reply_text("❌ Не удалось найти текст этой песни на Genius.")
        else:
            update.message.reply_text("❌ Не удалось скачать аудио (YouTube заблокировал запрос 403).")
    else:
        update.message.reply_text("❌ Не удалось найти песню по вашему запросу.")

def handle_audio(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id
    if user_id in BANNED_USERS:
        return

    file = update.message.audio or update.message.voice or update.message.document
    if not file:
        return
    file_id = file.file_id
    new_file = context.bot.get_file(file_id)
    filename = f"{file_id}.mp3"
    new_file.download(filename)
    orig_name = getattr(file, 'file_name', None)
    user_audio[user_id] = (filename, orig_name)
    update.message.reply_text("Укажи промежуток для вырезки в секундах, например: 68-73")

def save_query_to_excel(user_id, username, query):
    file_exists = os.path.exists(LOG_FILE)
    if not file_exists:
        wb = Workbook()
        ws = wb.active
        ws.title = "История поиска"
        ws.append(["User ID", "Username/Name", "Date & Time", "Query"])
    else:
        try:
            wb = load_workbook(LOG_FILE)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.append(["User ID", "Username/Name", "Date & Time", "Query"])

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([user_id, username, current_time, query])
    wb.save(LOG_FILE)

def search_youtube_track(query):
    ydl_opts = {
        'quiet': True, 
        'extract_flat': True, 
        'skip_download': True,
        'extractor_args': {'youtube': {'player_client': ['android', 'web']}},
        'http_headers': {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36'
        }
    }
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        search = f"ytsearch1:{query} музыка"
        try:
            result = ydl.extract_info(search, download=False)
            if 'entries' in result and result['entries']:
                entry = result['entries'][0]
                title = entry.get('title')
                url = f"https://www.youtube.com/watch?v={entry.get('id')}"
                return (title, url)
        except Exception:
            return None
    return None

def download_audio(url):
    ydl_opts = {
        'format': 'bestaudio/best',
        'outtmpl': '%(title)s.%(ext)s',
        'quiet': True,
        'extractor_args': {'youtube': {'player_client': ['android', 'web']}},
        'http_headers': {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36'
        },
        'postprocessors': [{
            'key': 'FFmpegExtractAudio',
            'preferredcodec': 'mp3',
            'preferredquality': '192',
        }],
    }
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        try:
            info = ydl.extract_info(url, download=True)
            filename = ydl.prepare_filename(info)
            filename = os.path.splitext(filename)[0] + '.mp3'
            return filename if os.path.exists(filename) else None
        except Exception as e:
            print(f"Download Error: {e}")
            return None

def clean_lyrics(lyrics):
    pattern = re.compile(r"(\[.*?\])", re.UNICODE)
    match = pattern.search(lyrics)
    if match:
        return lyrics[match.start():].strip()
    lines = lyrics.splitlines()
    for i, line in enumerate(lines):
        if len(line.strip()) > 20:
            return "\n".join(lines[i:]).strip()
    return lyrics.strip()

def get_lyrics_from_genius_api(query):
    headers = {"Authorization": f"Bearer {GENIUS_TOKEN}"}
    search_url = "https://api.genius.com/search"
    params = {"q": query}
    try:
        response = requests.get(search_url, params=params, headers=headers)
        data = response.json()
        hits = data.get("response", {}).get("hits", [])
        if not hits:
            return None
        song_info = hits[0]["result"]
        song_title = song_info["full_title"]
        song_url = song_info["url"]
        page = requests.get(song_url)
        soup = BeautifulSoup(page.text, "html.parser")
        lyrics_blocks = soup.find_all("div", {"data-lyrics-container": "true"})
        lyrics = "\n".join([block.get_text(separator="\n") for block in lyrics_blocks])
        lyrics = "\n".join([line.strip() for line in lyrics.splitlines() if line.strip()])
        if not lyrics:
            return None
        lyrics = clean_lyrics(lyrics)
        if len(lyrics) > 4000:
            lyrics = lyrics[:4000] + "\n...\n(текст обрезан)"
        return f"{song_title}\n\n{lyrics}"
    except Exception as e:
        print(f"Error getting lyrics: {e}")
        return None

def main():
    updater = Updater(API_TOKEN, use_context=True)
    dp = updater.dispatcher

    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(MessageHandler(Filters.audio | Filters.voice | Filters.document.audio, handle_audio))
    dp.add_handler(MessageHandler(Filters.text & ~Filters.command, handle_message))

    updater.start_polling()
    updater.idle()

if __name__ == '__main__':
    main()